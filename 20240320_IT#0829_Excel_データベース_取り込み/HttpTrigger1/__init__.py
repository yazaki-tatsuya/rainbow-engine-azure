import logging
import azure.functions as func
import openpyxl
import pyodbc
from io import BytesIO
from azure.storage.blob import BlobServiceClient
# 環境変数読み込み
import HttpTrigger1.env as env

def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Python HTTP trigger function processed a request.')

    # Blob名をリクエストから取得
    blob_name = req.params.get('blobname')
    if not blob_name:
        try:
            req_body = req.get_json()
        except ValueError:
            pass
        else:
            blob_name = req_body.get('blobname')

    if not blob_name:
        return func.HttpResponse(
            "Please pass the blobname on the query string or in the request body",
            status_code=400
        )

    # Blob Storageの接続文字列とコンテナ名
    blob_connection_string = env.get_env_variable("STORAGE_CONNECTION_STRING")
    container_name = env.get_env_variable("STORAGE_CONTAINER_NAME")
    blob_service_client = BlobServiceClient.from_connection_string(blob_connection_string)
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)

    # Blobをメモリ上に読み込む
    blob_data = blob_client.download_blob().readall()

    # Excelファイルを開く
    workbook = openpyxl.load_workbook(filename=BytesIO(blob_data))
    sheet = workbook.active

    # SQL Serverの接続情報
    server = env.get_env_variable("SQL_SERVER_NAME")
    database = env.get_env_variable("SQL_DATABASE_NAME")
    username = env.get_env_variable("SQL_DATABASE_USER")
    password = env.get_env_variable("SQL_DATABASE_PASS")
    driver= '{ODBC Driver 17 for SQL Server}'
    cnxn = pyodbc.connect('DRIVER=' + driver + ';SERVER=' + server + ';PORT=1433;DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
    cursor = cnxn.cursor()

    # Excelファイルからデータを読み取り、SQL Serverに登録
    try:

        # テーブルのデータを全て削除
        delete_query = "DELETE FROM TEST_EXCEL_IMPORT"
        cursor.execute(delete_query)

        # iter_rows()のmin_rowを3に設定して1行目（ヘッダー）をスキップ
        for row in sheet.iter_rows(min_row=3, values_only=True):
            # 空行をチェック（全てのセルがNoneである行をスキップ）
            if all(cell is None for cell in row):
                continue
            insert_query = "INSERT INTO TEST_EXCEL_IMPORT (カラム1, カラム2, カラム3, カラム4_1, カラム4_2, カラム4_3, カラム4_4) VALUES (?, ?, ?, ?, ?, ?, ?)"
            cursor.execute(insert_query, row[0], row[1], row[2], row[3], row[4], row[5], row[6])
        cnxn.commit()
    except Exception as e:
        cursor.close()
        cnxn.close()
        return func.HttpResponse(
            f"Failed to insert data into Azure SQL database. Error: {str(e)}",
            status_code=500
        )

    # 接続を閉じる
    cursor.close()
    cnxn.close()

    # 処理が成功した場合のメッセージを返す
    return func.HttpResponse(f"Excel data successfully inserted into Azure SQL database for blob {blob_name}.", status_code=200)